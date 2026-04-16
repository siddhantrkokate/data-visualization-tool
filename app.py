import random
import smtplib
import mysql.connector
from flask import Flask, render_template, request, redirect, session, flash, jsonify, make_response, url_for, send_file
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import subprocess
import os
import uuid
import json
import requests
import pandas as pd
from werkzeug.utils import secure_filename
from functools import wraps
import sys

app = Flask(__name__)
app.secret_key = "super_secret_key_change_this"

# Global config — loaded from config.json at startup
config = {}

UPLOAD_FOLDER = "files"
ALLOWED_EXTENSIONS = {"csv", "xlsx"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

AI_MODEL   = "gpt-oss-120b"
AI_API_URL = "https://cloud.olakrutrim.com/v1/chat/completions"

CONFIG_FILE = "config.json"
SQL_FILE    = "dvt_platform_backup.sql"


# =========================
# ERROR HANDLER
# =========================
@app.errorhandler(Exception)
def handle_exception(e):
    from flask import request as freq
    if freq.path.startswith("/project") and freq.method == "POST":
        return jsonify({"success": False, "error": str(e)}), 500
    raise e


# =========================
# DATABASE — reads from config
# =========================
def get_db():
    return mysql.connector.connect(
        host=config.get("db_host", "localhost"),
        user=config.get("db_user", "root"),
        password=config.get("db_pass", ""),
        database="dvt_platform"
    )


# =========================
# HELPERS
# =========================
def generate_otp():
    return str(random.randint(100000, 999999))


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_api_key():
    return config.get("krutrim_api_key", "")


def ai_post(prompt, temperature=0.2, max_tokens=1024, timeout=30):
    """Call the Krutrim AI API and return the raw text response."""
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer " + get_api_key()
    }
    payload = {
        "model": AI_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": temperature,
        "max_tokens": max_tokens
    }
    response = requests.post(AI_API_URL, headers=headers, json=payload, timeout=timeout)
    response.raise_for_status()
    return response.json()["choices"][0]["message"]["content"].strip()


def strip_code_fences(text):
    """Remove markdown code fences from AI response."""
    if text.startswith("```"):
        lines = text.splitlines()
        text = "\n".join(lines[1:])
        text = text.rsplit("```", 1)[0].strip()
    return text


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "email" not in session:
            return redirect(url_for("auth_page"))
        return f(*args, **kwargs)
    return decorated_function


# =========================
# EMAIL
# =========================
def send_email(receiver_email, otp):
    sender_email = config.get("sender_email", "")
    app_password = config.get("email_pass", "")

    plain = f"""DVT — Verification Code\n\nYour one-time code: {otp}\n\nExpires in 5 minutes.\n— DVT Platform"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8"/></head>
<body style="margin:0;padding:0;background:#f5f5f5;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="padding:48px 0;background:#f5f5f5;">
  <tr><td align="center">
    <table width="100%" cellpadding="0" cellspacing="0" style="max-width:480px;">
      <tr><td style="background:#0a0a0a;border-radius:16px 16px 0 0;padding:22px 32px;">
        <span style="font-size:15px;font-weight:700;color:#fff;">DVT</span>
      </td></tr>
      <tr><td style="background:#fff;padding:40px 32px;border:1px solid #e8e8e8;">
        <h1 style="margin:0 0 8px;font-size:22px;font-weight:700;color:#0a0a0a;text-align:center;">Verify your identity</h1>
        <p style="margin:0 0 32px;font-size:14px;color:#737373;text-align:center;">Use the code below — expires in <strong>5 minutes</strong>.</p>
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:32px;">
          <tr><td align="center">
            <div style="background:#0a0a0a;border-radius:14px;padding:20px 48px;display:inline-block;">
              <span style="font-size:38px;font-weight:700;color:#fff;letter-spacing:14px;font-family:'Courier New',monospace;">{otp}</span>
            </div>
          </td></tr>
        </table>
        <div style="background:#f5f5f5;border:1px solid #ebebeb;border-radius:10px;padding:14px 16px;">
          <p style="margin:0;font-size:12px;color:#737373;">
            <strong style="color:#0a0a0a;">Didn't request this?</strong><br/>You can safely ignore this email.
          </p>
        </div>
      </td></tr>
      <tr><td style="background:#fafafa;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 16px 16px;padding:18px 32px;">
        <p style="margin:0;font-size:11px;color:#b8b8b8;">© 2026 DVT Platform</p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body>
</html>"""

    msg = MIMEMultipart("alternative")
    msg["From"]    = sender_email
    msg["To"]      = receiver_email
    msg["Subject"] = "Your DVT Verification Code"
    msg.attach(MIMEText(plain, "plain"))
    msg.attach(MIMEText(html, "html"))

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, app_password)
        server.sendmail(sender_email, receiver_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print("Email Error:", e)
        return False


# =========================
# AUTH
# =========================
@app.route("/auth")
def auth_page():
    if "email" in session:
        return redirect(url_for("home_page"))
    resp = make_response(render_template("auth.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"]  = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/send-otp", methods=["POST"])
def send_otp():
    email = request.form.get("email", "").strip()
    if not email:
        flash("Email is required", "error")
        return redirect(url_for("auth_page"))

    db     = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT email FROM users WHERE email=%s", (email,))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO users (email) VALUES (%s)", (email,))
        db.commit()

    otp = generate_otp()
    cursor.execute("DELETE FROM otp_veri WHERE email=%s", (email,))
    cursor.execute("INSERT INTO otp_veri (email, otp) VALUES (%s, %s)", (email, otp))
    db.commit()
    cursor.close()
    db.close()

    if send_email(email, otp):
        session["temp_email"] = email
        flash("Verification code sent successfully.", "success")
    else:
        flash("Failed to send OTP. Try again.", "error")

    return redirect(url_for("auth_page"))


@app.route("/verify-otp", methods=["POST"])
def verify_otp():
    otp_entered = request.form.get("otp", "").strip()
    email       = session.get("temp_email")

    if not email:
        flash("Session expired. Please try again.", "error")
        return redirect(url_for("auth_page"))

    db     = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT otp FROM otp_veri WHERE email=%s AND otp=%s", (email, otp_entered))
    result = cursor.fetchone()

    if result:
        cursor.execute("DELETE FROM otp_veri WHERE email=%s", (email,))
        db.commit()
        cursor.close()
        db.close()
        session.pop("temp_email", None)
        session["email"] = email
        return redirect(url_for("home_page"))
    else:
        cursor.close()
        db.close()
        flash("Invalid OTP. Please try again.", "error")
        return redirect(url_for("auth_page"))


@app.route("/resend-otp", methods=["POST"])
def resend_otp():
    data  = request.get_json()
    email = data.get("email", "").strip()
    if not email:
        return jsonify({"success": False})

    otp    = generate_otp()
    db     = get_db()
    cursor = db.cursor()
    cursor.execute("DELETE FROM otp_veri WHERE email=%s", (email,))
    cursor.execute("INSERT INTO otp_veri (email, otp) VALUES (%s, %s)", (email, otp))
    db.commit()
    cursor.close()
    db.close()

    return jsonify({"success": send_email(email, otp)})


@app.route("/check-session")
def check_session():
    return jsonify({"logged_in": "email" in session})


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("auth_page"))


# =========================
# HOME
# =========================
@app.route("/")
@login_required
def home_page():
    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT project_id, project_name, project_category,
               project_date, project_time, file_name
        FROM projects WHERE email=%s ORDER BY created_at DESC
    """, (session["email"],))
    projects = cursor.fetchall()
    cursor.close()
    db.close()
    return render_template("home.html", projects=projects, user_email=session["email"])


# =========================
# CREATE PROJECT
# =========================
@app.route("/create")
@login_required
def create_page():
    return render_template("create.html", user_email=session["email"])


@app.route("/create-project", methods=["POST"])
@login_required
def create_project():
    project_name     = request.form.get("project_name", "").strip()
    project_category = request.form.get("project_category", "")
    file             = request.files.get("file")

    if len(project_name) < 3:
        flash("Project name must be at least 3 characters.", "error")
        return redirect(url_for("create_page"))

    if not file or file.filename == "":
        flash("Please upload a file.", "error")
        return redirect(url_for("create_page"))

    if not allowed_file(file.filename):
        flash("Only CSV and XLSX files are allowed.", "error")
        return redirect(url_for("create_page"))

    ext             = secure_filename(file.filename).rsplit(".", 1)[1].lower()
    unique_filename = f"{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(app.config["UPLOAD_FOLDER"], unique_filename))

    now = datetime.now()
    db     = get_db()
    cursor = db.cursor()
    cursor.execute("""
        INSERT INTO projects (email, project_name, project_category, project_date, project_time, file_name)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (session["email"], project_name, project_category, now.date(), now.strftime("%H:%M:%S"), unique_filename))
    db.commit()
    cursor.close()
    db.close()

    flash("Project created successfully!", "success")
    return redirect(url_for("create_page", created="1"))


# =========================
# PROJECT PAGE
# =========================
@app.route("/project/<int:project_id>")
@login_required
def project_page(project_id):
    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()
    cursor.close()
    db.close()

    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home_page"))

    file_path = os.path.join("files", project["file_name"])
    if not os.path.exists(file_path):
        flash("File not found.", "error")
        return redirect(url_for("home_page"))

    try:
        df = pd.read_csv(file_path) if file_path.endswith(".csv") else pd.read_excel(file_path)

        # Detect if last row is a score row
        scores_exist    = False
        existing_scores = {}
        if len(df) > 0:
            try:
                last_vals = df.iloc[-1].apply(pd.to_numeric, errors="coerce")
                if last_vals.notna().all() and last_vals.between(0, 100).all():
                    scores_exist    = True
                    existing_scores = {k: int(v) for k, v in last_vals.items()}
            except Exception:
                pass

        data_df         = df.iloc[:-1] if scores_exist else df
        total_records   = len(data_df)
        invalid_records = int(data_df.isnull().any(axis=1).sum())
        valid_records   = total_records - invalid_records
        warning_records = int(data_df.duplicated().sum())
        preview_data    = data_df.head(20).to_dict(orient="records")
        columns         = df.columns.tolist()

    except Exception as e:
        print("File Read Error:", e)
        flash("Error reading file.", "error")
        return redirect(url_for("home_page"))

    return render_template(
        "project.html",
        project=project,
        total_records=total_records,
        valid_records=valid_records,
        invalid_records=invalid_records,
        warning_records=warning_records,
        columns=columns,
        preview_data=preview_data,
        scores_exist=scores_exist,
        existing_scores=existing_scores,
        user_email=session["email"]
    )


# =========================
# AI COLUMN SCORING
# =========================
@app.route("/project/<int:project_id>/score-columns", methods=["POST"])
@login_required
def score_columns(project_id):
    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()
    cursor.close()
    db.close()

    if not project:
        return jsonify({"success": False, "error": "Project not found."}), 404

    file_path = os.path.join("files", project["file_name"])
    if not os.path.exists(file_path):
        return jsonify({"success": False, "error": "File not found."}), 404

    try:
        df = pd.read_csv(file_path) if file_path.endswith(".csv") else pd.read_excel(file_path)
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read file: {e}"}), 500

    columns  = df.columns.tolist()
    category = project.get("project_category", "general data")

    prompt = f"""You are a data quality expert. Assign a relevance score (integer 0–100) to each column based on importance for a '{category}' dataset.

Scoring guide:
- 75-100: Core, critical columns
- 50-74: Important but secondary
- 25-49: Marginally relevant or generic
- 0-24: Irrelevant or unclear

Columns: {", ".join(columns)}

Respond ONLY with a valid JSON object. No explanation, no markdown.
Example: {{"column_name": 85, "other_column": 42}}"""

    try:
        ai_text          = strip_code_fences(ai_post(prompt, temperature=0.2))
        scores           = json.loads(ai_text)
        validated_scores = {col: max(0, min(100, int(scores.get(col, 50)))) for col in columns}
    except requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "AI request timed out."}), 504
    except requests.exceptions.RequestException as e:
        return jsonify({"success": False, "error": f"AI API error: {e}"}), 502
    except (json.JSONDecodeError, ValueError) as e:
        return jsonify({"success": False, "error": f"Could not parse AI response: {e}"}), 500

    # Save score row back to file
    try:
        df_fresh = pd.read_csv(file_path) if file_path.endswith(".csv") else pd.read_excel(file_path)

        # Remove old score row if present
        if len(df_fresh) > 0:
            try:
                last_vals = df_fresh.iloc[-1].apply(pd.to_numeric, errors="coerce")
                if last_vals.notna().all() and last_vals.between(0, 100).all():
                    df_fresh = df_fresh.iloc[:-1]
            except Exception:
                pass

        score_row  = pd.DataFrame([{col: int(validated_scores.get(col, 0)) for col in df_fresh.columns}])
        updated_df = pd.concat([df_fresh, score_row], ignore_index=True)

        if file_path.endswith(".csv"):
            updated_df.to_csv(file_path, index=False)
        else:
            updated_df.to_excel(file_path, index=False)
    except Exception as e:
        print("File save error:", e)

    return jsonify({"success": True, "scores": validated_scores, "columns": columns})


# =========================
# GENERATE AI PROMPTS
# =========================
@app.route("/project/<int:project_id>/generate-prompts", methods=["POST"])
@login_required
def generate_prompts(project_id):
    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()
    cursor.close()
    db.close()

    if not project:
        return jsonify({"success": False, "error": "Project not found."}), 404

    data    = request.get_json()
    columns = data.get("columns", [])
    if not columns:
        return jsonify({"success": False, "error": "No columns provided."}), 400

    category = project.get("project_category", "general data")
    col_list = ", ".join(columns)

    prompt = f"""You are a data validation expert. For each column from a '{category}' dataset, write a concise validation prompt (max 80 words, comma-separated conditions).

Columns: {col_list}

Respond ONLY with a valid JSON object. No markdown, no explanation.
Example: {{"column_name": "must be non-empty, must be valid email, max 100 characters"}}"""

    try:
        ai_text = strip_code_fences(ai_post(prompt, temperature=0.3, max_tokens=2048))
        prompts = json.loads(ai_text)
        cleaned = {}
        for col in columns:
            raw   = prompts.get(col, "Must be non-empty, must have valid format, no null values.")
            words = raw.split()
            cleaned[col] = " ".join(words[:80])
        return jsonify({"success": True, "prompts": cleaned})
    except requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "Request timed out."}), 504
    except requests.exceptions.RequestException as e:
        return jsonify({"success": False, "error": f"API error: {e}"}), 502
    except (json.JSONDecodeError, KeyError) as e:
        return jsonify({"success": False, "error": f"Could not parse AI response: {e}"}), 500


# =========================
# START PROCESSING
# =========================
@app.route("/project/<int:project_id>/start-processing", methods=["POST"])
@login_required
def start_processing(project_id):
    import threading

    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()
    cursor.close()
    db.close()

    if not project:
        return jsonify({"success": False, "error": "Project not found."}), 404

    data       = request.get_json()
    col_config = data.get("config", [])
    if not col_config:
        return jsonify({"success": False, "error": "No column config provided."}), 400

    file_name   = project["file_name"]
    file_path   = os.path.join("files", file_name)
    base_name   = os.path.splitext(file_name)[0]
    runner_dir  = "runner"
    os.makedirs(runner_dir, exist_ok=True)

    script_path = os.path.join(runner_dir, base_name + ".py")
    status_path = os.path.join(runner_dir, base_name + "_status.json")
    output_path = os.path.join("files", base_name + ".xlsx")

    col_rules_str = "\n".join(
        f'Column "{c["column"]}": {c["prompt"]}' for c in col_config
    )

    # ─── FIXED AI PROMPT ──────────────────────────────────────────────────────
    # The Efficiency_Score formula is hardcoded here so the AI cannot deviate.
    # score = (columns_passed / total_columns) * 100  →  deterministic & accurate
    # ──────────────────────────────────────────────────────────────────────────
    ai_prompt = (
        "You are an expert Python developer. Write a complete standalone Python script.\n\n"
        f"INPUT FILE  : {file_path}\n"
        f"OUTPUT FILE : {output_path}\n"
        f"STATUS FILE : {status_path}\n\n"
        "VALIDATION RULES PER COLUMN:\n"
        f"{col_rules_str}\n\n"
        "WRITE THE SCRIPT USING EXACTLY THIS STRUCTURE:\n\n"
        "import pandas as pd\n"
        "import json, os, re, shutil\n"
        "from openpyxl import Workbook\n"
        "from openpyxl.styles import PatternFill\n\n"
        f"INPUT_FILE  = r'{file_path}'\n"
        f"OUTPUT_FILE = r'{output_path}'\n"
        f"STATUS_FILE = r'{status_path}'\n\n"
        "RED_FILL   = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')\n"
        "GREEN_FILL = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')\n"
        "AMBER_FILL = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')\n\n"
        "def write_status(status, progress, total, processed, invalid_rows_count, logs, error=None):\n"
        "    tmp = STATUS_FILE + '.tmp'\n"
        "    with open(tmp, 'w') as f:\n"
        "        json.dump({\n"
        "            'status': status, 'progress': progress, 'total_rows': total,\n"
        "            'processed_rows': processed, 'invalid_rows': invalid_rows_count,\n"
        "            'logs': logs, 'error': error\n"
        "        }, f)\n"
        "    shutil.move(tmp, STATUS_FILE)\n\n"
        "# ── Per-column validation functions ───────────────────────────────────\n"
        "# For EACH column listed in VALIDATION RULES, write:\n"
        "#   def validate_<safe_col_name>(value) -> tuple[bool, str]\n"
        "# Return (True, '') on pass, (False, 'short reason') on fail.\n"
        "# Use re, isinstance, str checks as appropriate.\n\n"
        "# ── Dispatcher ────────────────────────────────────────────────────────\n"
        "# def validate_column(col_name: str, value) -> tuple[bool, str]:\n"
        "#     Map col_name to the correct validate_* function.\n"
        "#     Return (True, '') for any unknown column (safe fallback).\n\n"
        "def main():\n"
        "    write_status('running', 0, 0, 0, 0, ['Reading file...'])\n\n"
        "    if INPUT_FILE.endswith('.csv'):\n"
        "        df = pd.read_csv(INPUT_FILE)\n"
        "    else:\n"
        "        df = pd.read_excel(INPUT_FILE)\n\n"
        "    # Drop trailing AI-score row if all values are numeric 0-100\n"
        "    if len(df) > 0:\n"
        "        last = df.iloc[-1].apply(pd.to_numeric, errors='coerce')\n"
        "        if last.notna().all() and last.between(0, 100).all():\n"
        "            df = df.iloc[:-1].reset_index(drop=True)\n\n"
        "    COLUMNS    = df.columns.tolist()\n"
        "    total_rows = len(df)\n"
        "    write_status('running', 0, total_rows, 0, 0, ['Starting validation...'])\n\n"
        "    # cell_pass[row_index] = list of bool, one per column\n"
        "    cell_pass    = []\n"
        "    cell_reasons = []  # parallel list of failure reasons\n\n"
        "    for i, row in df.iterrows():\n"
        "        row_pass    = []\n"
        "        row_reasons = []\n"
        "        for col in COLUMNS:\n"
        "            passed, reason = validate_column(col, row[col])\n"
        "            row_pass.append(passed)\n"
        "            row_reasons.append(reason)\n"
        "        cell_pass.append(row_pass)\n"
        "        cell_reasons.append(row_reasons)\n"
        "        if (i + 1) % 50 == 0 or (i + 1) == total_rows:\n"
        "            pct = int((i + 1) / total_rows * 85)\n"
        "            write_status('running', pct, total_rows, i + 1, 0,\n"
        "                         [f'Validated {i + 1}/{total_rows} rows...'])\n\n"
        "    # ── EFFICIENCY SCORE — FIXED FORMULA — DO NOT CHANGE ─────────────\n"
        "    # For every row: Efficiency_Score = (passed_columns / total_columns) * 100\n"
        "    # Rounded to 2 decimal places.\n"
        "    # 100.00 = every column passed  (HIGH ACCURACY)\n"
        "    #   0.00 = no column passed     (LOW ACCURACY)\n"
        "    # This formula is deterministic and must not be altered.\n"
        "    efficiency_scores = []\n"
        "    for row_pass in cell_pass:\n"
        "        if len(COLUMNS) == 0:\n"
        "            efficiency_scores.append(0.0)\n"
        "        else:\n"
        "            score = round(sum(row_pass) / len(COLUMNS) * 100, 2)\n"
        "            efficiency_scores.append(score)\n\n"
        "    df['Efficiency_Score'] = efficiency_scores\n\n"
        "    # ── Write XLSX with colour-coded cells ────────────────────────────\n"
        "    wb = Workbook()\n"
        "    ws = wb.active\n"
        "    ws.title = 'Validated Data'\n\n"
        "    all_cols = COLUMNS + ['Efficiency_Score']\n"
        "    for c_idx, col in enumerate(all_cols, 1):\n"
        "        ws.cell(row=1, column=c_idx, value=col)\n\n"
        "    invalid_rows_count = 0\n"
        "    for r_idx, (row_pass_list, score) in enumerate(zip(cell_pass, efficiency_scores), 2):\n"
        "        row_has_failure = not all(row_pass_list)\n"
        "        if row_has_failure:\n"
        "            invalid_rows_count += 1\n"
        "        for c_idx, (col, passed) in enumerate(zip(COLUMNS, row_pass_list), 1):\n"
        "            val  = df.at[r_idx - 2, col]\n"
        "            cell = ws.cell(row=r_idx, column=c_idx, value=val)\n"
        "            if not passed:\n"
        "                cell.fill = RED_FILL  # failed cell → red\n"
        "        # Efficiency_Score cell: green ≥75, amber 50-74, red <50\n"
        "        eff_cell = ws.cell(row=r_idx, column=len(COLUMNS) + 1, value=score)\n"
        "        if score >= 75:\n"
        "            eff_cell.fill = GREEN_FILL\n"
        "        elif score >= 50:\n"
        "            eff_cell.fill = AMBER_FILL\n"
        "        else:\n"
        "            eff_cell.fill = RED_FILL\n\n"
        "    wb.save(OUTPUT_FILE)\n\n"
        "    high_count = sum(1 for s in efficiency_scores if s >= 75)\n"
        "    low_count  = total_rows - high_count\n"
        "    high_pct   = round(high_count / total_rows * 100, 1) if total_rows else 0\n"
        "    low_pct    = round(low_count  / total_rows * 100, 1) if total_rows else 0\n"
        "    avg_score  = round(sum(efficiency_scores) / total_rows, 2) if total_rows else 0\n\n"
        "    write_status(\n"
        "        'complete', 100, total_rows, total_rows, invalid_rows_count,\n"
        "        [\n"
        "            f'Processing complete.',\n"
        "            f'Total rows       : {total_rows}',\n"
        "            f'High accuracy (≥75%) : {high_count} rows ({high_pct}%)',\n"
        "            f'Low  accuracy (<75%) : {low_count} rows ({low_pct}%)',\n"
        "            f'Average efficiency   : {avg_score}%',\n"
        "        ]\n"
        "    )\n\n"
        "if __name__ == '__main__':\n"
        "    main()\n\n"
        "CRITICAL RULES — MUST FOLLOW:\n"
        "1. Write validate_<safe_col_name>(value) for EVERY column in VALIDATION RULES.\n"
        "   - safe_col_name = col name lowercased, spaces/special chars replaced with underscore.\n"
        "2. write validate_column(col_name, value) dispatcher that routes to each function.\n"
        "3. The Efficiency_Score formula above is LOCKED. Do NOT change it in any way.\n"
        "4. Do NOT add extra scoring logic, weights, or multipliers.\n"
        "5. Output ONLY the complete Python script. NO markdown fences. NO explanation.\n"
    )

    try:
        script_code = strip_code_fences(ai_post(ai_prompt, temperature=0.1, max_tokens=8192, timeout=90))
    except Exception as e:
        return jsonify({"success": False, "error": f"AI generation failed: {e}"}), 500

    with open(script_path, "w", encoding="utf-8") as f:
        f.write(script_code)

    # Write initial status
    import shutil
    tmp = status_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump({"status": "running", "progress": 0, "total_rows": 0,
                   "processed_rows": 0, "invalid_rows": 0,
                   "logs": ["Script generated. Execution starting..."], "error": None}, f)
    shutil.move(tmp, status_path)

    def run_script():
        try:
            subprocess.run(["python", script_path], timeout=600, check=False)
        except Exception as e:
            import shutil as _sh
            etmp = status_path + ".tmp"
            with open(etmp, "w") as sf:
                json.dump({"status": "error", "progress": 0, "total_rows": 0,
                           "processed_rows": 0, "invalid_rows": 0,
                           "logs": [f"Execution error: {e}"], "error": str(e)}, sf)
            _sh.move(etmp, status_path)

    threading.Thread(target=run_script, daemon=True).start()

    return jsonify({"success": True, "script_path": script_path, "job_id": base_name})


# =========================
# PROCESSING STATUS
# =========================
@app.route("/project/<int:project_id>/processing-status/<job_id>")
@login_required
def processing_status(project_id, job_id):
    import time
    safe_id     = secure_filename(job_id)
    status_path = os.path.join("runner", safe_id + "_status.json")

    if not os.path.exists(status_path):
        return jsonify({"status": "pending", "progress": 0, "logs": ["Waiting to start..."],
                        "invalid_rows": 0, "total_rows": 0, "processed_rows": 0})

    for _ in range(5):
        try:
            with open(status_path, "r", encoding="utf-8") as f:
                data = f.read().strip()
            if data:
                return jsonify(json.loads(data))
        except Exception:
            pass
        time.sleep(0.1)

    return jsonify({"status": "running", "progress": 0, "logs": ["Processing..."],
                    "invalid_rows": 0, "total_rows": 0, "processed_rows": 0})


# =========================
# VIEW GENERATED SCRIPT
# =========================
@app.route("/project/<int:project_id>/get-script/<job_id>")
@login_required
def get_script(project_id, job_id):
    script_path = os.path.join("runner", secure_filename(job_id) + ".py")
    if not os.path.exists(script_path):
        return jsonify({"success": False, "error": "Script not found."}), 404
    with open(script_path, "r", encoding="utf-8") as f:
        return jsonify({"success": True, "code": f.read()})


# =========================
# FULL FILE DATA  ← returns ALL rows + accuracy stats
# =========================
@app.route("/project/<int:project_id>/full-data")
@login_required
def full_data(project_id):
    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()
    cursor.close()
    db.close()

    if not project:
        return jsonify({"success": False, "error": "Project not found"}), 404

    base_name = os.path.splitext(project["file_name"])[0]
    xlsx_path = os.path.join("files", base_name + ".xlsx")
    csv_path  = os.path.join("files", project["file_name"])
    file_path = xlsx_path if os.path.exists(xlsx_path) else csv_path

    if not os.path.exists(file_path):
        return jsonify({"success": False, "error": "File not found"}), 404

    try:
        if file_path.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb        = load_workbook(file_path)
            ws        = wb.active
            headers   = []
            rows      = []
            red_cells = []

            for r_idx, row in enumerate(ws.iter_rows()):
                if r_idx == 0:
                    headers = [str(c.value) if c.value is not None else "" for c in row]
                    continue
                row_data = []
                for c_idx, cell in enumerate(row):
                    row_data.append("" if cell.value is None else cell.value)
                    fill = cell.fill
                    if fill and fill.fill_type == "solid" and fill.fgColor:
                        rgb = getattr(fill.fgColor, "rgb", "")
                        if rgb in ("FFFF0000", "FF0000"):
                            red_cells.append({"row": r_idx - 1, "col": c_idx})
                rows.append(row_data)
        else:
            df      = pd.read_csv(file_path)
            headers = list(df.columns)
            rows    = [["" if (isinstance(v, float) and pd.isna(v)) else v for v in r]
                       for r in df.values.tolist()]
            red_cells = []

        # ── Accuracy stats derived from Efficiency_Score column ──────────────
        accuracy_stats = {}
        if "Efficiency_Score" in headers:
            eff_idx = headers.index("Efficiency_Score")
            scores  = []
            for row in rows:
                try:
                    val = row[eff_idx]
                    if val != "":
                        scores.append(float(val))
                except (ValueError, TypeError, IndexError):
                    pass

            if scores:
                total      = len(scores)
                high_rows  = [s for s in scores if s >= 75]
                mid_rows   = [s for s in scores if 50 <= s < 75]
                low_rows   = [s for s in scores if s < 50]
                accuracy_stats = {
                    "total_rows"          : total,
                    "high_accuracy_count" : len(high_rows),
                    "mid_accuracy_count"  : len(mid_rows),
                    "low_accuracy_count"  : len(low_rows),
                    "high_accuracy_pct"   : round(len(high_rows) / total * 100, 1),
                    "mid_accuracy_pct"    : round(len(mid_rows)  / total * 100, 1),
                    "low_accuracy_pct"    : round(len(low_rows)  / total * 100, 1),
                    "avg_efficiency"      : round(sum(scores) / total, 2),
                    "max_efficiency"      : round(max(scores), 2),
                    "min_efficiency"      : round(min(scores), 2),
                }

        return jsonify({
            "success"        : True,
            "headers"        : headers,
            "rows"           : rows,       # ALL rows — nothing hidden
            "red_cells"      : red_cells,
            "total"          : len(rows),
            "accuracy_stats" : accuracy_stats,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


# =========================
# DOWNLOAD FILTERED DATA
# =========================
@app.route("/project/<int:project_id>/download-filtered", methods=["POST"])
@login_required
def download_filtered(project_id):
    from io import BytesIO
    from openpyxl import Workbook as XLWorkbook

    data    = request.get_json()
    headers = data.get("headers", [])
    rows    = data.get("rows", [])

    if not headers or not rows:
        return jsonify({"success": False, "error": "No data"}), 400

    wb = XLWorkbook()
    ws = wb.active
    ws.title = "Filtered Data"
    for c_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=c_idx, value=str(h))
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="filtered_data.xlsx")


# =========================
# DOWNLOAD HIGH ACCURACY ROWS ONLY
# =========================
@app.route("/project/<int:project_id>/download-high-accuracy", methods=["GET"])
@login_required
def download_high_accuracy(project_id):
    """Downloads only rows with Efficiency_Score >= 75."""
    from io import BytesIO
    from openpyxl import Workbook as XLWorkbook

    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()
    cursor.close()
    db.close()

    if not project:
        return jsonify({"success": False, "error": "Project not found"}), 404

    base_name = os.path.splitext(project["file_name"])[0]
    xlsx_path = os.path.join("files", base_name + ".xlsx")
    csv_path  = os.path.join("files", project["file_name"])
    file_path = xlsx_path if os.path.exists(xlsx_path) else csv_path

    if not os.path.exists(file_path):
        return jsonify({"success": False, "error": "File not found"}), 404

    try:
        if file_path.endswith(".xlsx"):
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)

        if "Efficiency_Score" not in df.columns:
            return jsonify({"success": False, "error": "Efficiency_Score column not found. Run processing first."}), 400

        df["Efficiency_Score"] = pd.to_numeric(df["Efficiency_Score"], errors="coerce").fillna(0)
        high_df = df[df["Efficiency_Score"] >= 75]

        wb = XLWorkbook()
        ws = wb.active
        ws.title = "High Accuracy"
        for c_idx, col in enumerate(high_df.columns, 1):
            ws.cell(row=1, column=c_idx, value=col)
        for r_idx, (_, row) in enumerate(high_df.iterrows(), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name="high_accuracy_rows.xlsx")
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# =========================
# STOP PROCESSING
# =========================
@app.route("/project/<int:project_id>/stop-processing/<job_id>", methods=["POST"])
@login_required
def stop_processing(project_id, job_id):
    import shutil
    status_path = os.path.join("runner", job_id + "_status.json")
    tmp_path    = status_path + ".tmp"
    try:
        current = {}
        if os.path.exists(status_path):
            with open(status_path) as f:
                current = json.load(f)
        current["status"] = "stopped"
        current["error"]  = "Stopped by user"
        with open(tmp_path, "w") as f:
            json.dump(current, f)
        shutil.move(tmp_path, status_path)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# =========================
# DELETE PROJECT
# =========================
@app.route("/delete-project/<int:project_id>", methods=["POST"])
@login_required
def delete_project(project_id):
    db     = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    project = cursor.fetchone()

    if not project:
        cursor.close()
        db.close()
        return jsonify({"success": False, "error": "Project not found."}), 404

    base_name = os.path.splitext(project["file_name"])[0]

    for fname in [project["file_name"], base_name + ".xlsx"]:
        path = os.path.join("files", fname)
        if os.path.exists(path):
            try: os.remove(path)
            except Exception as e: print(f"Could not delete {path}: {e}")

    for suffix in [".py", "_status.json"]:
        path = os.path.join("runner", base_name + suffix)
        if os.path.exists(path):
            try: os.remove(path)
            except Exception as e: print(f"Could not delete {path}: {e}")

    cursor.execute("DELETE FROM projects WHERE project_id=%s AND email=%s",
                   (project_id, session["email"]))
    db.commit()
    cursor.close()
    db.close()

    return jsonify({"success": True})


# =========================
# SETUP HELPERS
# =========================
def _print_banner():
    print("\n  ╔══════════════════════════════════════╗")
    print("  ║          DVT Platform Setup          ║")
    print("  ╚══════════════════════════════════════╝\n")

def _ok(msg):   print(f"  ✔  {msg}")
def _err(msg):  print(f"  ✘  {msg}")
def _info(msg): print(f"  ·  {msg}")
def _ask(prompt): return input(f"  ▸  {prompt}: ").strip()


def setup_mysql():
    print("  ── MySQL Database ─────────────────────────────")
    while True:
        host     = _ask("Host [localhost]") or "localhost"
        user     = _ask("User")
        password = _ask("Password")
        try:
            conn   = mysql.connector.connect(host=host, user=user, password=password)
            cursor = conn.cursor()
            cursor.execute("CREATE DATABASE IF NOT EXISTS dvt_platform")
            conn.commit()
            cursor.execute("USE dvt_platform")
            cursor.execute("SHOW TABLES")
            tables = [r[0] for r in cursor.fetchall()]
            cursor.close()
            conn.close()

            if "projects" in tables and "users" in tables:
                _ok("Tables already exist — skipping SQL import.")
            else:
                if not os.path.exists(SQL_FILE):
                    _err(f"'{SQL_FILE}' not found. Place it next to app.py.")
                    sys.exit(1)
                _info(f"Importing '{SQL_FILE}'…")
                with open(SQL_FILE, "r", encoding="utf-8") as f:
                    sql_content = f.read()
                result = subprocess.run(
                    ["mysql", "-h", host, "-u", user, f"-p{password}", "dvt_platform"],
                    input=sql_content, text=True, capture_output=True
                )
                if result.returncode != 0:
                    _err(f"SQL import failed: {result.stderr.strip()}")
                    sys.exit(1)
                _ok("Database imported successfully.")
            _ok(f"Connected to MySQL at {host}.")
            return host, user, password
        except mysql.connector.Error as e:
            _err(f"MySQL connection failed: {e}")
            _info("Check your credentials and try again.\n")


def setup_smtp():
    print("\n  ── Gmail SMTP ─────────────────────────────────")
    _info("You need a Gmail App Password (not your regular password).")
    _info("Generate one at: myaccount.google.com/apppasswords\n")
    while True:
        sender   = _ask("Sender Gmail address")
        app_pass = _ask("Gmail App Password")
        if not sender or "@" not in sender:
            _err("Enter a valid Gmail address."); continue
        if not app_pass:
            _err("App password cannot be empty."); continue
        try:
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(sender, app_pass)
            server.quit()
            _ok(f"SMTP authenticated as {sender}.")
            return sender, app_pass
        except smtplib.SMTPAuthenticationError:
            _err("Authentication failed — check your credentials.\n")
        except Exception as e:
            _err(f"SMTP error: {e}\n")


def setup_krutrim():
    print("\n  ── Krutrim AI API ─────────────────────────────")
    _info("Get your API key from: cloud.olakrutrim.com\n")
    while True:
        api_key = _ask("Krutrim API Key")
        if not api_key:
            _err("API key cannot be empty."); continue
        _info("Verifying API key…")
        try:
            resp = requests.post(
                AI_API_URL,
                headers={"Content-Type": "application/json",
                         "Authorization": f"Bearer {api_key}"},
                json={"model": AI_MODEL, "messages": [{"role": "user", "content": "ping"}],
                      "max_tokens": 5},
                timeout=15
            )
            if resp.status_code == 401:
                _err("Invalid API key. Try again.\n"); continue
            _ok(f"API reachable (status {resp.status_code}) — key accepted.")
            return api_key
        except requests.exceptions.Timeout:
            _info("Verification timed out — accepting key.")
            return api_key
        except Exception as e:
            _info(f"Could not verify ({e}) — accepting key.")
            return api_key


def load_or_setup():
    global config

    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            # Quick DB ping
            conn = mysql.connector.connect(
                host=cfg["db_host"], user=cfg["db_user"],
                password=cfg["db_pass"], database="dvt_platform"
            )
            conn.close()
            if not cfg.get("sender_email"):
                raise KeyError("sender_email missing")
            config.update(cfg)
            _ok("Config loaded successfully.")
            return
        except (mysql.connector.Error, KeyError, json.JSONDecodeError):
            print("\n  ⚠  Config invalid or DB unreachable — running setup…\n")

    _print_banner()
    db_host, db_user, db_pass = setup_mysql()
    sender_email, email_pass  = setup_smtp()
    api_key                   = setup_krutrim()

    cfg = {
        "db_host": db_host, "db_user": db_user, "db_pass": db_pass,
        "sender_email": sender_email, "email_pass": email_pass,
        "krutrim_api_key": api_key
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)

    config.update(cfg)

    print("\n  ╔══════════════════════════════════════╗")
    print("  ║   Setup complete — starting DVT…     ║")
    print("  ╚══════════════════════════════════════╝\n")


# =========================
# ENTRY POINT
# =========================
if __name__ == "__main__":
    os.makedirs("files",  exist_ok=True)
    os.makedirs("runner", exist_ok=True)
    load_or_setup()
    app.run(debug=True)
